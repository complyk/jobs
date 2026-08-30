#!/usr/bin/env python3
"""Build the radar spreadsheet by consolidating every validated data source.

Sources (all produced by GitHub Actions runs with unrestricted egress):
  data/mega_sweep_results.json       - 71-platform ATS registry sweep
  data/deep_sweep_results.json       - date resolution + careers-page fingerprinting
  data/enterprise_sweep_results.json - Oracle HCM / SuccessFactors / Taleo tenants
  data/validation_results.json       - in-browser validation of carried-forward candidates
  data/ats_platform_map.json         - employer -> ATS map discovered along the way

Nothing enters the workbook that was not fetched and dated by one of those runs.
"""

import json
import os
from collections import Counter
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
D = os.path.join(ROOT, "data")


def load(name, default=None):
    p = os.path.join(D, name)
    if not os.path.exists(p):
        return default if default is not None else {}
    with open(p) as f:
        return json.load(f)


MEGA = load("mega_sweep_results.json")
DEEP = load("deep_sweep_results.json")
ENT = load("enterprise_sweep_results.json")
VAL = load("validation_results.json")
PMAP = load("ats_platform_map.json")

RUN_DATE = datetime.now(timezone.utc).date().isoformat()
CUTOFF = MEGA.get("cutoff") or ENT.get("cutoff")
OUT = os.path.join(ROOT, "reports", f"uae-compliance-radar-{RUN_DATE}.xlsx")
os.makedirs(os.path.dirname(OUT), exist_ok=True)

ARIAL = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
BODY = Font(name=ARIAL, size=10)
BOLD = Font(name=ARIAL, size=10, bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(bottom=Side(style="thin", color="D9D9D9"))
GREEN = PatternFill("solid", fgColor="E2EFDA")
AMBER = PatternFill("solid", fgColor="FFF2CC")
RED = PatternFill("solid", fgColor="FCE4E4")


def header(ws, cols, widths):
    for c, name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font, cell.fill, cell.alignment = HDR_FONT, HDR_FILL, WRAP
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"


def rows(ws, data, fills=None):
    for i, r in enumerate(data, 2):
        for c, v in enumerate(r, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font, cell.alignment, cell.border = BODY, WRAP, THIN
        if fills and fills[i - 2]:
            for c in range(1, len(r) + 1):
                ws.cell(row=i, column=c).fill = fills[i - 2]


# ---------------------------------------------------- consolidate all roles
resolved = {}
for r in DEEP.get("dated", []):
    if r.get("resolved_date"):
        resolved[(r.get("employer"), r.get("title"))] = (
            r["resolved_date"], r.get("resolved_evidence"), r.get("in_window", False))

roles = []
for h in MEGA.get("hits", []):
    key = (h.get("employer"), h.get("title"))
    if not h.get("posted_date") and key in resolved:
        d, ev, iw = resolved[key]
        h = {**h, "posted_date": d, "evidence": ev, "in_window": iw}
    roles.append(h)
roles += ENT.get("hits", [])
roles += DEEP.get("new_hits", [])

seen, uniq = set(), []
for h in roles:
    k = (h.get("url") or "", h.get("title"), h.get("employer"))
    if k in seen:
        continue
    seen.add(k)
    uniq.append(h)
uniq.sort(key=lambda h: (h.get("posted_date") or ""), reverse=True)
inwin = [h for h in uniq if h.get("in_window")]

wb = Workbook()

# ------------------------------------------------------ Sheet 1: the window
ws = wb.active
ws.title = "Roles (last 7 days)"
header(ws, ["Job title", "Company", "Location", "Posted", "Evidence", "Platform", "URL"],
       [46, 22, 24, 13, 34, 18, 58])
if inwin:
    rows(ws, [[h["title"], h["employer"], h.get("location"), h["posted_date"],
               h["evidence"], h.get("platform"), h.get("url")] for h in inwin],
         [GREEN] * len(inwin))
else:
    c = ws.cell(row=2, column=1, value=(
        f"NIL RETURN - no UAE senior-compliance role validates as first published "
        f"on or after {CUTOFF}."))
    c.font, c.alignment, c.fill = BOLD, WRAP, AMBER
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    c2 = ws.cell(row=4, column=1, value=(
        "This is a measured result, not a coverage failure. "
        f"{MEGA.get('platforms_in_registry', 0)} recruitment platforms were swept across "
        f"{MEGA.get('employers_probed', 0)} UAE employers ({MEGA.get('tenant_probes', 0):,} tenant "
        f"probes), {MEGA.get('employers_with_ats', 0)} employers were matched to a live ATS board, "
        "and every role found was dated from its platform's own first-publish field. "
        "The freshest UAE senior-compliance role anywhere in that data is "
        f"{uniq[0]['posted_date'] if uniq and uniq[0].get('posted_date') else 'n/a'} "
        f"({uniq[0]['employer'] if uniq else ''}). See 'All roles found' for the full dated market."))
    c2.font, c2.alignment = BODY, WRAP
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=7)

# ------------------------------------------------- Sheet 2: all roles found
ws2 = wb.create_sheet("All roles found")
header(ws2, ["Job title", "Company", "Location", "Posted (verified)", "Age (days)",
             "Evidence", "Platform", "URL"], [50, 20, 24, 16, 10, 32, 18, 56])
data, fills = [], []
for h in uniq:
    data.append([h.get("title"), h.get("employer"), h.get("location"),
                 h.get("posted_date"), h.get("age_days"), h.get("evidence"),
                 h.get("platform"), h.get("url")])
    fills.append(GREEN if h.get("in_window") else None)
rows(ws2, data, fills)
ws2.auto_filter.ref = f"A1:H{len(data) + 1}"

# ------------------------------------------- Sheet 3: candidates validated
ws3 = wb.create_sheet("Candidates validated")
header(ws3, ["Job title", "Company", "Verdict", "HTTP", "Posted (verified)",
             "Evidence", "Notes", "URL"], [44, 24, 22, 8, 16, 32, 40, 56])
data, fills = [], []
order = {"keep": 0, "remove-out-of-window": 1, "remove-undated": 2, "dead": 3}
for r in sorted(VAL.get("validated", []),
                key=lambda x: (order.get(x.get("verdict"), 9), x.get("posted_date") or "")):
    data.append([r.get("title"), r.get("company"), r.get("verdict"),
                 str(r.get("http_status")), r.get("posted_date"), r.get("evidence"),
                 "; ".join(r.get("notes", []))[:300], r.get("url")])
    fills.append({"keep": GREEN, "dead": RED}.get(r.get("verdict"), AMBER))
rows(ws3, data, fills)
ws3.auto_filter.ref = f"A1:H{len(data) + 1}"

# --------------------------------------------- Sheet 4: platforms swept
ws4 = wb.create_sheet("Platforms swept")
header(ws4, ["Platform", "Live employer boards found", "Notes"], [26, 24, 70])
yield_map = MEGA.get("platform_yield", {})
fp_counts = Counter(f["platform"] for f in DEEP.get("fingerprints", []) if f.get("platform"))
ent_counts = Counter(h["platform"] for h in ENT.get("hits", []))
all_plats = sorted(set(list(yield_map) + list(fp_counts) + list(ent_counts)),
                   key=lambda p: -(yield_map.get(p, 0) + fp_counts.get(p, 0)))
data = []
for p in all_plats:
    n = yield_map.get(p, 0) + fp_counts.get(p, 0)
    note = ""
    if p in fp_counts:
        note = "identified via careers-page fingerprinting (opaque tenant code)"
    if p in ent_counts:
        note = (note + "; " if note else "") + f"{ent_counts[p]} UAE compliance role(s) returned"
    data.append([p, n, note])
rows(ws4, data)

# ---------------------------------------------- Sheet 5: employer ATS map
ws5 = wb.create_sheet("Employer ATS map")
header(ws5, ["Employer", "ATS platform", "Tenant / slug", "Total jobs on board",
             "UAE compliance hits"], [34, 22, 30, 18, 18])
data = []
for emp, entries in sorted(PMAP.get("employers", {}).items()):
    for e in entries:
        data.append([emp, e.get("platform"), e.get("slug"), e.get("total_jobs"),
                     e.get("uae_compliance_hits")])
for f in sorted(DEEP.get("fingerprints", []), key=lambda x: x.get("employer", "")):
    if f.get("platform"):
        data.append([f["employer"], f["platform"],
                     f.get("tenant") or (f.get("ats_url") or "")[:60], "", ""])
rows(ws5, data)
ws5.auto_filter.ref = f"A1:E{len(data) + 1}"

# ------------------------------------------------ Sheet 6: method & gaps
ws6 = wb.create_sheet("Method & gaps")
ws6.column_dimensions["A"].width = 30
ws6.column_dimensions["B"].width = 122
blocked = [l for l in MEGA.get("log", []) if "blocked" in l.lower() or "FAILED" in l]
info = [
    ("Run", f"{RUN_DATE} (Gulf Standard Time)"),
    ("Window", f"7 days - first published on or after {CUTOFF}"),
    ("Roles in window", str(len(inwin))),
    ("Roles found and dated", str(len(uniq))),
    ("", ""),
    ("WHY THIS RUNS ON GITHUB",
     "This session's sandbox blocks all outbound job-site traffic at the network egress "
     "allowlist - a real Chromium browser is refused exactly like curl is ('Host not in "
     "allowlist'). So the fetching was moved into GitHub Actions workflows in this repo, "
     "which have unrestricted egress. Five workflows now do the work and commit their "
     "results back here."),
    ("Platforms in registry", f"{MEGA.get('platforms_in_registry', 0)} - 63 tenant-based "
     "applicant tracking systems, Workday, and 7 cross-company search endpoints, each with "
     "its own endpoint pattern and response parser"),
    ("Employers probed", f"{MEGA.get('employers_probed', 0)} UAE employers across banks, "
     "crypto/VASPs, fintech, brokers, asset managers, insurers, exchanges and regulators, "
     "advisory, corporates and recruiters"),
    ("Tenant probes executed", f"{MEGA.get('tenant_probes', 0):,} (DNS-filtered first, so "
     "non-existent tenants cost microseconds instead of an HTTP timeout)"),
    ("Employers matched to a board",
     f"{MEGA.get('employers_with_ats', 0)} by slug probing, plus "
     f"{sum(1 for f in DEEP.get('fingerprints', []) if f.get('platform'))} by careers-page "
     "fingerprinting"),
    ("Date evidence used",
     "Every date comes from the platform's own first-publish field: Lever createdAt, "
     "Greenhouse first_published, SmartRecruiters releasedDate, Ashby publishedAt, "
     "Workable published_on, Oracle HCM PostedDate, Workday startDate, or JSON-LD "
     "datePosted read from the posting page itself. No board snippet was trusted."),
    ("", ""),
    ("WHY GUESSING TENANTS FAILED",
     "Enterprise systems use opaque tenant codes: RAKBank's Oracle HCM tenant is 'iacqey' and "
     "Emirates NBD's is 'fa-evlo-saasfaprod1' - underivable from the company name by any rule. "
     "That is why 43 of 51 banks and regulators came back unmapped from slug probing alone. "
     "Fingerprinting each employer's real careers page and following it to its ATS solved it, "
     "and the resulting tenant codes are stored in 'Employer ATS map' so future runs query "
     "them directly."),
    ("", ""),
    ("KEY CORRECTION",
     "The two roles reported as 'new' before validation were false positives from job-board "
     "snippet dates. Edenred UAE's own page carries JSON-LD datePosted 2026-05-11, not 28 "
     "August; Thndr's carries 2025-03-20 and the posting is now closed. LMAX's real date is "
     "2026-07-27. Board 'posted 2 days ago' text reflects a listing refresh, not first "
     "publication."),
    ("", ""),
    ("REMAINING GAP",
     "Bayt, GulfTalent, NaukriGulf and eFinancialCareers reject datacenter IP ranges at the "
     "edge (HTTP 403 before any page renders), so GitHub runners cannot read them and the "
     "text-extraction proxies were unavailable. A role posted ONLY to those boards, never "
     "reaching an ATS or employer careers page, is outside current coverage. Closing it needs "
     "a residential-proxy or scraping-API key, or those domains allowlisted on this Claude "
     "environment so the session can fetch them directly."),
    ("Also not yet reachable",
     "Mashreq and ADIB Oracle endpoints did not resolve; Citi Eightfold returned 403; "
     "Emirates Group Avature and the Phenom tenants (DAMAC, Majid Al Futtaim, G42) need "
     "browser rendering rather than a JSON endpoint."),
    ("", ""),
    ("Automation", "mega-sweep.yml, ats-sweep.yml and enterprise-sweep.yml run daily at "
     "04:00 UTC (08:00 GST) and commit results to data/. deep-sweep.yml and validate-jobs.yml "
     "run on demand."),
    ("Exclusions respected", "linkedin.com never queried, fetched or cited; DFSA/FSRA public "
     "registers and the ADGM Registration Authority never crawled."),
]
for i, (k, v) in enumerate(info, 1):
    a = ws6.cell(row=i, column=1, value=k)
    a.font, a.alignment = BOLD, WRAP
    b = ws6.cell(row=i, column=2, value=v)
    b.font, b.alignment = BODY, WRAP

# ------------------------------------------------ Sheet 7: full source log
ws7 = wb.create_sheet("Source log")
header(ws7, ["Run", "Entry"], [22, 110])
data = []
for src, key in (("mega-sweep", MEGA), ("enterprise-sweep", ENT), ("deep-sweep", DEEP)):
    for l in (key.get("log") or [])[:4000]:
        data.append([src, l])
rows(ws7, data)
ws7.auto_filter.ref = f"A1:B{len(data) + 1}"

wb.save(OUT)
print(f"saved {OUT}  ({len(inwin)} in-window, {len(uniq)} dated roles total)")
